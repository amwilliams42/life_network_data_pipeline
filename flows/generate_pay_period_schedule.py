"""
Pay Period Schedule Workbook Builder

Pure functions for building Excel workbooks for schedule reports.
No database access - that's handled by the Prefect tasks in schedule_reports.py.

This module provides:
- create_summary_sheet: Summary tab with employee hours breakdown
- create_daily_sheet: Daily shift list tab
- create_comparison_summary_sheet: Comparison summary with actual vs scheduled
- create_comparison_daily_sheet: Daily comparison with variance
- Helper functions for hour calculations
"""

from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# Constants
# =============================================================================

REGION_TO_SOURCE = {
    'il': 'il',
    'mi': 'mi',
    'tn_memphis': 'tn',
    'tn_nashville': 'tn',
}

REGION_DISPLAY_NAMES = {
    'il': 'Illinois',
    'mi': 'Michigan',
    'tn_memphis': 'Memphis',
    'tn_nashville': 'Nashville',
}

# Styling
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
OPEN_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
TRAINING_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# =============================================================================
# Helper Functions
# =============================================================================

def set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=True):
    """Helper to set cell value and styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = THIN_BORDER
    return cell


def get_calendar_week_key(shift_date):
    """Get the Sunday start of the calendar week containing shift_date."""
    # Python weekday(): Monday=0, Sunday=6
    # We want Sunday=0, so: (weekday + 1) % 7
    days_since_sunday = (shift_date.weekday() + 1) % 7
    sunday = shift_date - timedelta(days=days_since_sunday)
    return sunday


def calculate_employee_hours(shifts, include_actual=False):
    """
    Calculate employee hours by pay period week and calendar week overtime.

    Returns dict keyed by employee name with:
    - week1_hours, week2_hours (pay period weeks)
    - total_ot (overtime per calendar week, >40 hours)
    - training_hours, field_hours, orientation_hours, special_event_hours
    - If include_actual: actual_hours, variance
    """
    employees = defaultdict(lambda: {
        'user_id': None,
        'week1_hours': 0,
        'week2_hours': 0,
        'training_hours': 0,
        'field_hours': 0,
        'orientation_hours': 0,
        'special_event_hours': 0,
        'calendar_weeks': defaultdict(float),
        'actual_hours': 0,
    })

    for shift in shifts:
        if shift['is_open']:
            continue

        name = shift['assigned_name']
        hours = float(shift['scheduled_hours'] or 0)
        actual = float(shift['actual_hours_worked'] or 0) if include_actual else 0
        pp_week = shift['pp_week']
        calendar_week = get_calendar_week_key(shift['shift_date'])

        emp = employees[name]
        emp['user_id'] = shift['user_id']

        # Pay period week hours
        if pp_week == 1:
            emp['week1_hours'] += hours
        else:
            emp['week2_hours'] += hours

        # Category hours
        if shift['is_training']:
            emp['training_hours'] += hours
        if shift['is_orientation']:
            emp['orientation_hours'] += hours
        elif shift['is_special_event']:
            emp['special_event_hours'] += hours
        elif shift['is_field_shift']:
            emp['field_hours'] += hours

        # Calendar week hours for OT calculation
        emp['calendar_weeks'][calendar_week] += hours

        # Actual hours
        if include_actual:
            emp['actual_hours'] += actual

    # Calculate overtime per calendar week
    for name, emp in employees.items():
        emp['total_ot'] = 0
        for week_start, week_hours in emp['calendar_weeks'].items():
            if week_hours > 40:
                emp['total_ot'] += week_hours - 40

        emp['total_hours'] = emp['week1_hours'] + emp['week2_hours']
        emp['variance'] = emp['actual_hours'] - emp['total_hours']

    return dict(employees)


def calculate_open_hours(shifts):
    """Calculate total open hours and open hours by category."""
    open_hours = {
        'total': 0,
        'field': 0,
        'orientation': 0,
        'special_event': 0,
        'training': 0,
        'week1': 0,
        'week2': 0,
    }

    for shift in shifts:
        if not shift['is_open']:
            continue

        hours = float(shift['scheduled_hours'] or 0)
        open_hours['total'] += hours

        if shift['pp_week'] == 1:
            open_hours['week1'] += hours
        else:
            open_hours['week2'] += hours

        if shift['is_training']:
            open_hours['training'] += hours
        if shift['is_orientation']:
            open_hours['orientation'] += hours
        elif shift['is_special_event']:
            open_hours['special_event'] += hours
        else:
            open_hours['field'] += hours

    return open_hours


# =============================================================================
# Summary Sheet Builders
# =============================================================================

def create_summary_sheet(wb, shifts, pp_number, pp_start, pp_end):
    """Create the summary sheet with employee hours breakdown."""
    ws = wb.active
    ws.title = "Summary"

    employee_hours = calculate_employee_hours(shifts)
    open_hours = calculate_open_hours(shifts)

    # Title
    row = 1
    ws.merge_cells(f'A{row}:I{row}')
    title_cell = ws.cell(row=row, column=1,
                         value=f"Pay Period {pp_number}: {pp_start.strftime('%b %d')} - {pp_end.strftime('%b %d, %Y')}")
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")
    row += 2

    # Column headers
    headers = [
        ("Employee", 25),
        ("Week 1 Hrs", 12),
        ("Week 2 Hrs", 12),
        ("Total Hrs", 12),
        ("OT Hrs", 10),
        ("Training Hrs", 13),
        ("Field Hrs", 11),
        ("Orientation", 12),
        ("Special Event", 13),
    ]

    for col, (header, width) in enumerate(headers, 1):
        set_cell(ws, row, col, header, font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal="center", vertical="center"))
        ws.column_dimensions[get_column_letter(col)].width = width
    row += 1

    # Employee rows
    sorted_employees = sorted(employee_hours.items(), key=lambda x: (x[0] or '').lower())

    totals = {
        'week1': 0, 'week2': 0, 'total': 0, 'ot': 0,
        'training': 0, 'field': 0, 'orientation': 0, 'special_event': 0
    }

    for name, emp in sorted_employees:
        values = [
            name,
            round(emp['week1_hours'], 1),
            round(emp['week2_hours'], 1),
            round(emp['total_hours'], 1),
            round(emp['total_ot'], 1) if emp['total_ot'] > 0 else "",
            round(emp['training_hours'], 1) if emp['training_hours'] > 0 else "",
            round(emp['field_hours'], 1) if emp['field_hours'] > 0 else "",
            round(emp['orientation_hours'], 1) if emp['orientation_hours'] > 0 else "",
            round(emp['special_event_hours'], 1) if emp['special_event_hours'] > 0 else "",
        ]

        for col, val in enumerate(values, 1):
            alignment = Alignment(horizontal="left" if col == 1 else "center")
            set_cell(ws, row, col, val, alignment=alignment)

        totals['week1'] += emp['week1_hours']
        totals['week2'] += emp['week2_hours']
        totals['total'] += emp['total_hours']
        totals['ot'] += emp['total_ot']
        totals['training'] += emp['training_hours']
        totals['field'] += emp['field_hours']
        totals['orientation'] += emp['orientation_hours']
        totals['special_event'] += emp['special_event_hours']
        row += 1

    # Open hours row
    open_values = [
        "OPEN",
        round(open_hours['week1'], 1),
        round(open_hours['week2'], 1),
        round(open_hours['total'], 1),
        "",  # No OT for open shifts
        round(open_hours['training'], 1) if open_hours['training'] > 0 else "",
        round(open_hours['field'], 1) if open_hours['field'] > 0 else "",
        round(open_hours['orientation'], 1) if open_hours['orientation'] > 0 else "",
        round(open_hours['special_event'], 1) if open_hours['special_event'] > 0 else "",
    ]

    for col, val in enumerate(open_values, 1):
        alignment = Alignment(horizontal="left" if col == 1 else "center")
        set_cell(ws, row, col, val, fill=OPEN_FILL, alignment=alignment)
    row += 1

    # Totals row
    total_values = [
        "TOTAL",
        round(totals['week1'] + open_hours['week1'], 1),
        round(totals['week2'] + open_hours['week2'], 1),
        round(totals['total'] + open_hours['total'], 1),
        round(totals['ot'], 1) if totals['ot'] > 0 else "",
        round(totals['training'] + open_hours['training'], 1) if (totals['training'] + open_hours['training']) > 0 else "",
        round(totals['field'] + open_hours['field'], 1) if (totals['field'] + open_hours['field']) > 0 else "",
        round(totals['orientation'] + open_hours['orientation'], 1) if (totals['orientation'] + open_hours['orientation']) > 0 else "",
        round(totals['special_event'] + open_hours['special_event'], 1) if (totals['special_event'] + open_hours['special_event']) > 0 else "",
    ]

    for col, val in enumerate(total_values, 1):
        alignment = Alignment(horizontal="left" if col == 1 else "center")
        set_cell(ws, row, col, val, font=Font(bold=True), fill=TOTAL_FILL, alignment=alignment)

    return ws


def create_comparison_summary_sheet(wb, shifts, pp_number, pp_start, pp_end):
    """Create the comparison summary sheet with scheduled vs actual hours."""
    ws = wb.active
    ws.title = "Summary"

    employee_hours = calculate_employee_hours(shifts, include_actual=True)
    open_hours = calculate_open_hours(shifts)

    # Title
    row = 1
    ws.merge_cells(f'A{row}:G{row}')
    title_cell = ws.cell(row=row, column=1,
                         value=f"Pay Period {pp_number} Comparison: {pp_start.strftime('%b %d')} - {pp_end.strftime('%b %d, %Y')}")
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")
    row += 2

    # Column headers
    headers = [
        ("Employee", 25),
        ("Scheduled Hrs", 14),
        ("Actual Hrs", 12),
        ("Variance", 12),
        ("OT Hrs", 10),
        ("Training Hrs", 13),
        ("Open Hrs (unfilled)", 16),
    ]

    for col, (header, width) in enumerate(headers, 1):
        set_cell(ws, row, col, header, font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal="center", vertical="center"))
        ws.column_dimensions[get_column_letter(col)].width = width
    row += 1

    # Employee rows
    sorted_employees = sorted(employee_hours.items(), key=lambda x: (x[0] or '').lower())

    totals = {
        'scheduled': 0, 'actual': 0, 'variance': 0, 'ot': 0, 'training': 0
    }

    for name, emp in sorted_employees:
        variance = round(emp['variance'], 1)
        values = [
            name,
            round(emp['total_hours'], 1),
            round(emp['actual_hours'], 1),
            variance if variance != 0 else "",
            round(emp['total_ot'], 1) if emp['total_ot'] > 0 else "",
            round(emp['training_hours'], 1) if emp['training_hours'] > 0 else "",
            "",  # Open hours column - empty for employees
        ]

        for col, val in enumerate(values, 1):
            alignment = Alignment(horizontal="left" if col == 1 else "center")
            fill = None
            if col == 4 and variance != 0:  # Variance column
                fill = NEGATIVE_FILL if variance < 0 else POSITIVE_FILL
            set_cell(ws, row, col, val, alignment=alignment, fill=fill)

        totals['scheduled'] += emp['total_hours']
        totals['actual'] += emp['actual_hours']
        totals['variance'] += emp['variance']
        totals['ot'] += emp['total_ot']
        totals['training'] += emp['training_hours']
        row += 1

    # Open hours row
    open_values = [
        "OPEN (unfilled)",
        round(open_hours['total'], 1),
        "",  # No actual for open
        "",  # No variance for open
        "",  # No OT for open
        round(open_hours['training'], 1) if open_hours['training'] > 0 else "",
        round(open_hours['total'], 1),
    ]

    for col, val in enumerate(open_values, 1):
        alignment = Alignment(horizontal="left" if col == 1 else "center")
        set_cell(ws, row, col, val, fill=OPEN_FILL, alignment=alignment)
    row += 1

    # Totals row
    total_variance = round(totals['variance'], 1)
    total_values = [
        "TOTAL",
        round(totals['scheduled'] + open_hours['total'], 1),
        round(totals['actual'], 1),
        total_variance if total_variance != 0 else "",
        round(totals['ot'], 1) if totals['ot'] > 0 else "",
        round(totals['training'] + open_hours['training'], 1) if (totals['training'] + open_hours['training']) > 0 else "",
        round(open_hours['total'], 1),
    ]

    for col, val in enumerate(total_values, 1):
        alignment = Alignment(horizontal="left" if col == 1 else "center")
        fill = TOTAL_FILL
        if col == 4 and total_variance != 0:
            fill = NEGATIVE_FILL if total_variance < 0 else POSITIVE_FILL
        set_cell(ws, row, col, val, font=Font(bold=True), fill=fill, alignment=alignment)

    return ws


# =============================================================================
# Daily Sheet Builders
# =============================================================================

def create_daily_sheet(wb, date, day_shifts):
    """Create a sheet for a specific day with shift details."""
    # Format: Mon-dd-mm (Excel doesn't allow / in sheet names)
    sheet_name = date.strftime("%a-%d-%m")
    ws = wb.create_sheet(title=sheet_name)

    # Title
    row = 1
    ws.merge_cells(f'A{row}:G{row}')
    title_cell = ws.cell(row=row, column=1, value=date.strftime("%A, %B %d, %Y"))
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    row += 2

    # Column headers
    headers = [
        ("Employee", 25),
        ("Unit", 18),
        ("Cost Center", 22),
        ("Start", 10),
        ("End", 10),
        ("Hours", 8),
        ("Training", 10),
    ]

    for col, (header, width) in enumerate(headers, 1):
        set_cell(ws, row, col, header, font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal="center", vertical="center"))
        ws.column_dimensions[get_column_letter(col)].width = width
    row += 1

    # Sort shifts: by start time, then unit, then employee
    sorted_shifts = sorted(day_shifts, key=lambda s: (
        s['shift_start'] or datetime.min.time(),
        s['unit_name'] or '',
        not s['is_open'],  # Open shifts after filled
        s['assigned_name'] or ''
    ))

    total_hours = 0
    open_hours = 0

    for shift in sorted_shifts:
        hours = float(shift['scheduled_hours'] or 0)
        total_hours += hours

        if shift['is_open']:
            open_hours += hours
            employee_name = "OPEN"
        else:
            employee_name = shift['assigned_name']

        values = [
            employee_name,
            shift['unit_name'],
            shift['cost_center_name'],
            shift['shift_start'].strftime('%I:%M %p') if shift['shift_start'] else "",
            shift['shift_end'].strftime('%I:%M %p') if shift['shift_end'] else "",
            round(hours, 1),
            "Yes" if shift['is_training'] else "",
        ]

        fill = None
        if shift['is_open']:
            fill = OPEN_FILL
        elif shift['is_training']:
            fill = TRAINING_FILL

        for col, val in enumerate(values, 1):
            alignment = Alignment(horizontal="left" if col in [1, 2, 3] else "center")
            set_cell(ws, row, col, val, fill=fill, alignment=alignment)
        row += 1

    # Summary row
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    set_cell(ws, row, 1, f"Total Hours: {round(total_hours, 1)}",
             font=Font(bold=True), fill=TOTAL_FILL,
             alignment=Alignment(horizontal="right"))
    set_cell(ws, row, 6, round(total_hours, 1), font=Font(bold=True), fill=TOTAL_FILL,
             alignment=Alignment(horizontal="center"))

    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    set_cell(ws, row, 1, f"Open Hours: {round(open_hours, 1)}",
             font=Font(bold=True), fill=OPEN_FILL,
             alignment=Alignment(horizontal="right"))
    set_cell(ws, row, 6, round(open_hours, 1), font=Font(bold=True), fill=OPEN_FILL,
             alignment=Alignment(horizontal="center"))

    return ws


def create_comparison_daily_sheet(wb, date, day_shifts):
    """Create a daily sheet for comparison report with actual hours."""
    sheet_name = date.strftime("%a-%d-%m")
    ws = wb.create_sheet(title=sheet_name)

    # Title
    row = 1
    ws.merge_cells(f'A{row}:H{row}')
    title_cell = ws.cell(row=row, column=1, value=date.strftime("%A, %B %d, %Y"))
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    row += 2

    # Column headers
    headers = [
        ("Employee", 25),
        ("Unit", 18),
        ("Cost Center", 22),
        ("Start", 10),
        ("End", 10),
        ("Scheduled", 10),
        ("Actual", 10),
        ("Variance", 10),
    ]

    for col, (header, width) in enumerate(headers, 1):
        set_cell(ws, row, col, header, font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal="center", vertical="center"))
        ws.column_dimensions[get_column_letter(col)].width = width
    row += 1

    # Sort shifts
    sorted_shifts = sorted(day_shifts, key=lambda s: (
        s['shift_start'] or datetime.min.time(),
        s['unit_name'] or '',
        not s['is_open'],
        s['assigned_name'] or ''
    ))

    total_scheduled = 0
    total_actual = 0
    open_hours = 0

    for shift in sorted_shifts:
        scheduled = float(shift['scheduled_hours'] or 0)
        actual = float(shift['actual_hours_worked'] or 0)
        variance = actual - scheduled
        total_scheduled += scheduled

        if shift['is_open']:
            open_hours += scheduled
            employee_name = "OPEN"
            actual = ""
            variance = ""
        else:
            employee_name = shift['assigned_name']
            total_actual += actual
            variance = round(variance, 1) if variance != 0 else ""

        values = [
            employee_name,
            shift['unit_name'],
            shift['cost_center_name'],
            shift['shift_start'].strftime('%I:%M %p') if shift['shift_start'] else "",
            shift['shift_end'].strftime('%I:%M %p') if shift['shift_end'] else "",
            round(scheduled, 1),
            round(actual, 1) if actual != "" else "",
            variance,
        ]

        fill = None
        if shift['is_open']:
            fill = OPEN_FILL
        elif shift['is_training']:
            fill = TRAINING_FILL

        for col, val in enumerate(values, 1):
            alignment = Alignment(horizontal="left" if col in [1, 2, 3] else "center")
            cell_fill = fill
            if col == 8 and variance != "" and variance != 0:
                cell_fill = NEGATIVE_FILL if variance < 0 else POSITIVE_FILL
            set_cell(ws, row, col, val, fill=cell_fill, alignment=alignment)
        row += 1

    # Summary rows
    row += 1
    total_variance = round(total_actual - (total_scheduled - open_hours), 1)

    ws.merge_cells(f'A{row}:E{row}')
    set_cell(ws, row, 1, "Totals", font=Font(bold=True), fill=TOTAL_FILL,
             alignment=Alignment(horizontal="right"))
    set_cell(ws, row, 6, round(total_scheduled, 1), font=Font(bold=True), fill=TOTAL_FILL,
             alignment=Alignment(horizontal="center"))
    set_cell(ws, row, 7, round(total_actual, 1), font=Font(bold=True), fill=TOTAL_FILL,
             alignment=Alignment(horizontal="center"))
    variance_fill = TOTAL_FILL
    if total_variance != 0:
        variance_fill = NEGATIVE_FILL if total_variance < 0 else POSITIVE_FILL
    set_cell(ws, row, 8, total_variance if total_variance != 0 else "",
             font=Font(bold=True), fill=variance_fill,
             alignment=Alignment(horizontal="center"))

    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    set_cell(ws, row, 1, f"Open Hours: {round(open_hours, 1)}",
             font=Font(bold=True), fill=OPEN_FILL,
             alignment=Alignment(horizontal="right"))
    set_cell(ws, row, 6, round(open_hours, 1), font=Font(bold=True), fill=OPEN_FILL,
             alignment=Alignment(horizontal="center"))

    return ws